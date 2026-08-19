import os
import shutil
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.utils import secure_filename
import csv
import io
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# --- IMPORTACIONES DE TU NÚCLEO (CORE) ---
from core.db_models import inicializar_base_datos, verificar_usuario, set_estado, get_estado, listar_usuarios, crear_usuario, eliminar_usuario, cambiar_clave_usuario
from core.inventory import obtener_inventario_tipos_documentales, obtener_modelos_conocidos, extension_permitida, borrar_todo_el_conocimiento, borrar_conocimiento_proceso, borrar_conocimiento_clase, descartar_subida_clase

# ==========================================
# 1. CONFIGURACIÓN DE CARPETAS Y LOGS
# ==========================================
base_dir = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(base_dir, '..', 'volumen_compartido', 'logs')
os.makedirs(log_dir, exist_ok=True) # Crea la carpeta si no existe

log_file = os.path.join(log_dir, 'intexus_auditor.log')

# Definimos cómo se verá cada línea (Fecha, Hora, Nivel de Alerta, Mensaje)
formato_log = logging.Formatter('[%(asctime)s] [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

# Manejador 1: Escribe en el archivo de texto (Máximo 5MB por archivo, guarda 3 históricos)
file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
file_handler.setFormatter(formato_log)

# Manejador 2: Imprime en la consola de Docker (para que sigamos viéndolo en vivo)
console_handler = logging.StreamHandler()
console_handler.setFormatter(formato_log)

# Configuramos el registrador principal
logger = logging.getLogger()
logger.setLevel(logging.INFO)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# Silenciamos un poco el ruido de Flask para que no ensucie nuestra bitácora
logging.getLogger('werkzeug').setLevel(logging.WARNING)
# --- FIN CONFIGURACIÓN DE LOGS ---

app = Flask(__name__)
csrf = CSRFProtect(app)

# ==========================================
# 2. INICIALIZACIÓN DE FLASK Y SEGURIDAD
# ==========================================

# Configurar Rate Limiter para prevenir ataques de fuerza bruta
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri=os.environ.get('CELERY_BROKER_URL', 'redis://localhost:6379/0')
)

# Obtenemos la llave inyectada por Docker (.env)
app.secret_key = os.environ.get('SECRET_KEY')

# Blindaje: Si no hay llave, detenemos el sistema por seguridad
if not app.secret_key:
    raise ValueError("🚨 ¡ALERTA DE SEGURIDAD! Falta la variable SECRET_KEY en el entorno.")

with app.app_context():
    inicializar_base_datos()

# ==========================================
# 3. MIDDLEWARE DE SEGURIDAD (Rutas Protegidas)
# ==========================================
def login_requerido(f):
    def wrap(*args, **kwargs):
        if 'usuario' not in session:
            flash("Por favor, inicia sesión primero.", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrap.__name__ = f.__name__
    return wrap

# ==========================================
# 4. RUTAS PRINCIPALES Y DE AUDITORÍA
# ==========================================
@app.route('/', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        rol = verificar_usuario(username, password)
        if rol:
            session['usuario'] = username
            session['rol'] = rol
            return redirect(url_for('dashboard'))
        else:
            flash("Usuario o contraseña incorrectos.", "danger")
            
    return render_template('login.html')

@app.route('/dashboard')
@login_requerido
def dashboard():
    modelos_conocidos = obtener_modelos_conocidos()
    return render_template('dashboard.html', usuario=session['usuario'], rol=session['rol'], modelos_conocidos=modelos_conocidos)

@app.route('/conocimiento')
@login_requerido
def conocimiento():
    modelos_conocidos = obtener_modelos_conocidos()
    return render_template(
        'conocimiento.html',
        usuario=session['usuario'],
        rol=session['rol'],
        modelos_conocidos=modelos_conocidos
    )

from core.auditor import procesar_lote_kofax_task
from core.db_models import obtener_estado_lote, obtener_resultados_lote, listar_lotes_auditoria
import uuid

@app.route('/api/estado_auditoria/<task_id>', methods=['GET'])
@limiter.exempt
@login_requerido
def api_estado_auditoria(task_id):
    estado = obtener_estado_lote(task_id)
    if not estado:
        return jsonify({"archivo": "No encontrado", "procesados": 0, "meta": 0, "estado": "error"})
    
    # Intentar obtener el progreso actual desde Celery si la tarea sigue activa
    from celery_app import celery
    task = celery.AsyncResult(task_id)
    archivo_actual = "Iniciando..."
    procesados_actual = estado['documentos_procesados']
    
    if task.state == 'PROGRESS':
        archivo_actual = task.info.get('archivo', 'Procesando...')
        procesados_actual = task.info.get('procesados', procesados_actual)
    elif estado['estado'] == 'completado':
        archivo_actual = "Completado"
    
    return jsonify({
        "archivo": archivo_actual,
        "procesados": procesados_actual,
        "meta": estado['total_documentos'],
        "estado": estado['estado']
    })

@app.route('/api/auditar_lote', methods=['POST'])
@login_requerido
def api_auditar_lote():
    # Salvaguarda: Evitar múltiples auditorías simultáneas
    lotes = listar_lotes_auditoria()
    for lote in lotes:
        if lote['estado'] == 'procesando':
            return jsonify({"error": "Ya hay una auditoría en curso. Por favor espera a que termine para lanzar otra."}), 400

    task_id = str(uuid.uuid4())
    procesar_lote_kofax_task.apply_async(args=[task_id], task_id=task_id)
    return jsonify({"status": "Lote enviado a procesamiento en segundo plano", "task_id": task_id})

@app.route('/api/auditoria_resultados/<task_id>', methods=['GET'])
@login_requerido
def api_auditoria_resultados(task_id):
    resultados = obtener_resultados_lote(task_id)
    estado = obtener_estado_lote(task_id)
    return jsonify({"resultados": resultados, "lote": estado})

@app.route('/api/historial_lotes', methods=['GET'])
@login_requerido
def api_historial_lotes():
    lotes = listar_lotes_auditoria()
    return jsonify(lotes)

@app.route('/api/descargar_reporte/<task_id>', methods=['GET'])
@login_requerido
def api_descargar_reporte(task_id):
    """Genera y descarga el reporte Excel (.xlsx) con el diseño original institucional."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    resultados = obtener_resultados_lote(task_id)
    estado = obtener_estado_lote(task_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte_Auditoria"

    # Encabezados exactos del formato original
    headers = ["Archivo TIF", "Proceso", "Clasificación Humana", "Clasificación IA", "Veredicto Final"]
    ws.append(headers)

    # Palette de estilos
    header_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")  # Verde Lima institucional
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_alignment = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Estilar fila de encabezado
    ws.row_dimensions[1].height = 26
    for col_num in range(1, 6):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    row_alt_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    row_white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    font_data = Font(name="Calibri", size=11)

    for i, r in enumerate(resultados, start=2):
        est = r.get('estado', 'danger')
        pred = r.get('prediccion', '')
        
        if est == 'success':
            veredicto = "MATCH PERFECTO  ✅ MATCH PERFECTO"
        elif est == 'warning':
            veredicto = "ALERTA - NO ENTRENADO  ⚠️ NO ENTRENADO"
        else:
            veredicto = f"ALERTA - IA DETECTÓ: {pred}  🚨 ALERTA - IA DETECTÓ: {pred}"

        archivo = r.get('archivo', '')
        proceso = r.get('subproceso', '')
        humano = r.get('esperado', '')
        ia = pred

        ws.append([archivo, proceso, humano, ia, veredicto])
        ws.row_dimensions[i].height = 22
        
        fill_actual = row_alt_fill if i % 2 == 0 else row_white_fill
        for col_num in range(1, 6):
            cell = ws.cell(row=i, column=col_num)
            cell.fill = fill_actual
            cell.font = font_data
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="center" if col_num == 2 else "left")

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 16)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Postgres devuelve datetime.datetime, SQLite devuelve string — manejamos ambos casos
    fecha_inicio_raw = estado.get('fecha_inicio') if estado else None
    if fecha_inicio_raw is None:
        fecha = 'sin_fecha'
    elif hasattr(fecha_inicio_raw, 'strftime'):
        fecha = fecha_inicio_raw.strftime('%Y-%m-%d')
    else:
        fecha = str(fecha_inicio_raw)[:10]
    nombre_archivo = f'Reporte_Auditoria_{fecha}_{task_id[:8]}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )

@app.route('/api/imagen/<path:nombre_archivo>', methods=['GET'])
@limiter.exempt
@login_requerido
def api_previsualizar_imagen(nombre_archivo):
    """Sirve una imagen del lote para previsualización en el navegador (convirtiendo TIF a JPEG si es necesario)."""
    lote_dir = '/volumen_compartido/lote_kofax'
    ruta = os.path.join(lote_dir, nombre_archivo)
    
    # Prevenir Path Traversal
    real_lote_dir = os.path.realpath(lote_dir)
    real_ruta = os.path.realpath(ruta)
    if not real_ruta.startswith(real_lote_dir):
        return jsonify({"error": "Acceso denegado. Ruta inválida."}), 403

    if not os.path.exists(real_ruta):
        return jsonify({"error": "Archivo no encontrado"}), 404
        
    try:
        from PIL import Image
        img = Image.open(real_ruta)
        # Convert to RGB (to handle CMYK, palettes, or TIFF specifics)
        if img.mode != 'RGB':
            img = img.convert('RGB')
            
        img_io = io.BytesIO()
        img.save(img_io, 'JPEG', quality=85)
        img_io.seek(0)
        return send_file(img_io, mimetype='image/jpeg')
    except Exception as e:
        logging.error(f"Error convirtiendo imagen {ruta}: {e}")
        return jsonify({"error": "No se pudo procesar la imagen"}), 500

@app.route('/api/discrepancias/<task_id>', methods=['GET'])
@login_requerido
def api_discrepancias(task_id):
    """Devuelve solo las filas con discrepancias para revisión manual."""
    resultados = obtener_resultados_lote(task_id)
    discrepancias = [r for r in resultados if r['estado'] == 'danger']
    return jsonify({"discrepancias": discrepancias})

@app.route('/api/aplicar_correcciones/<task_id>', methods=['POST'])
@login_requerido
def api_aplicar_correcciones(task_id):
    """Aplica solo las correcciones aprobadas por el usuario al archivo Indice_*.txt."""
    import glob
    data = request.get_json()
    correcciones_aprobadas = data.get('correcciones', [])  # Lista de {linea_indice, nuevo_valor}
    
    if not correcciones_aprobadas:
        return jsonify({"error": "No se recibieron correcciones para aplicar."}), 400
    
    lote_dir = '/volumen_compartido/lote_kofax'
    archivos_indice = [
        f for f in glob.glob(os.path.join(lote_dir, '*.[tT][xX][tT]'))
        if os.path.basename(f).lower().startswith('indice_')
    ]
    if not archivos_indice:
        return jsonify({"error": "No se encontró ningún archivo de índice ('Indice_*.txt' o 'indice_*.txt')."}), 404
    
    indice_path = archivos_indice[0]
    
    # Crear backup del índice original antes de aplicar las correcciones finales
    import shutil
    import time
    from datetime import datetime
    from werkzeug.utils import secure_filename
    
    backup_dir = '/volumen_compartido/backups_indices'
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    indice_nombre_base = os.path.basename(indice_path)
    indice_backup_path = os.path.join(backup_dir, f"{indice_nombre_base}.modificado.{timestamp}.bak")
    shutil.copy2(indice_path, indice_backup_path)
    logging.info(f"Respaldo pre-correcciones creado en: {indice_backup_path}")
    
    # Leer todas las líneas del archivo original
    with open(indice_path, 'r', encoding='utf-8', errors='ignore') as f:
        lineas = f.readlines()
    
    import shutil
    import time
    from werkzeug.utils import secure_filename
    
    # Crear un diccionario de correcciones aprobadas {linea: nuevo_valor}
    mapa_correcciones = {int(c['linea_indice']): c['nuevo_valor'] for c in correcciones_aprobadas}
    
    # Aplicar los cambios
    cambios_aplicados = 0
    imagenes_copiadas = 0
    dataset_base = '/volumen_compartido/dataset_entrenamiento'
    
    for num_linea, nuevo_valor in mapa_correcciones.items():
        if 1 <= num_linea <= len(lineas):
            linea_original = lineas[num_linea - 1]
            # Parseamos la línea como CSV para modificar solo el campo 14 (Tipo Documental)
            reader = csv.reader(io.StringIO(linea_original))
            for partes in reader:
                if len(partes) >= 16:
                    # Copiar la imagen al dataset de entrenamiento para aprendizaje continuo
                    archivo_imagen = partes[15].strip()
                    matriz = partes[13].strip().upper()[:2]
                    subproceso = partes[8].strip().upper()
                    nuevo_valor_seguro = secure_filename(nuevo_valor).replace("_", " ")
                    
                    if not (archivo_imagen.lower().endswith('.tif') or archivo_imagen.lower().endswith('.pdf') or archivo_imagen.lower().endswith('.jpg')):
                        archivo_imagen += '.TIF'
                        
                    ruta_origen = os.path.join(lote_dir, archivo_imagen)
                        
                    if os.path.exists(ruta_origen):
                        ruta_destino_dir = os.path.join(dataset_base, matriz, subproceso, nuevo_valor_seguro)
                        os.makedirs(ruta_destino_dir, exist_ok=True)
                        ruta_destino = os.path.join(ruta_destino_dir, f"{int(time.time())}_{archivo_imagen}")
                        try:
                            if not os.path.exists(ruta_destino):
                                shutil.copy2(ruta_origen, ruta_destino)
                                imagenes_copiadas += 1
                        except Exception as e:
                            logging.error(f"Error copiando {ruta_origen} a {ruta_destino}: {e}")
                            
                    partes[14] = nuevo_valor
                    # Reconstruimos la línea
                    output_line = io.StringIO()
                    writer = csv.writer(output_line)
                    writer.writerow(partes)
                    lineas[num_linea - 1] = output_line.getvalue()
                    cambios_aplicados += 1
    
    # Truco para evadir errores de permisos: borrar el archivo viejo (la carpeta sí tiene permisos 777)
    # y crear uno nuevo en su lugar, evitando el PermissionError al intentar sobreescribirlo.
    try:
        if os.path.exists(indice_path):
            os.remove(indice_path)
    except Exception as e:
        logging.warning(f"No se pudo eliminar {indice_path} previamente: {e}")
    
    # Escribir de vuelta (creará un archivo nuevo con permisos correctos del contenedor)
    with open(indice_path, 'w', encoding='utf-8', newline='') as f:
        f.writelines(lineas)
        
    # Verificar umbral de reentrenamiento POR CLASE específica
    try:
        if imagenes_copiadas > 0:
            UMBRAL = 5
            clases_listas = []
            # Recorrer el dataset pendiente y contar por clase (Matriz/Subproceso/TipoDocumental)
            for root, dirs, files in os.walk(dataset_base):
                if 'processed' in root.split(os.sep):
                    continue
                archivos_clase = [f for f in files if not f.endswith('.txt')]
                if len(archivos_clase) >= UMBRAL:
                    clases_listas.append(root)
            
            if clases_listas:
                set_estado('progreso_entrenamiento', '0')
                set_estado('entrenamiento', 'PROCESANDO')
                logging.info(f"🚀 ¡Umbral alcanzado! Clases listas para aprender: {clases_listas}. Disparando reentrenamiento.")
    except Exception as e:
        logging.error(f"Error comprobando umbral: {e}")
    
    logging.info(f"[AUDITOR] {session['usuario']} aplicó {cambios_aplicados} correcciones al índice.")
    return jsonify({"mensaje": f"Se aplicaron {cambios_aplicados} correcciones al archivo índice.", "cambios": cambios_aplicados})

# ==========================================
# 5. RUTAS DE ADMINISTRACIÓN Y ENTRENAMIENTO
# ==========================================
@app.route('/admin')
@login_requerido
def admin():
    if session['rol'] not in ('admin', 'superadmin'):
        flash("Acceso denegado. Módulo exclusivo para Administradores.", "danger")
        return redirect(url_for('dashboard'))

    inventario = obtener_inventario_tipos_documentales()
    modelos_conocidos = obtener_modelos_conocidos()

    return render_template(
        'admin.html',
        usuario=session['usuario'],
        rol=session['rol'],
        inventario=inventario,
        modelos_conocidos=modelos_conocidos
    )

@app.route('/admin/subir_documentos', methods=['POST'])
@login_requerido
def subir_documentos():
    if session['rol'] not in ('admin', 'superadmin'):
        return redirect(url_for('dashboard'))

    matriz = secure_filename(request.form.get('matriz', '').strip())
    subproceso_raw = request.form.get('subproceso', '').strip().upper()
    clase_doc = request.form.get('clase_documento', '').strip()
    archivos = request.files.getlist('archivos')

    # Separar subprocesos por coma y limpiarlos
    lista_subprocesos = [secure_filename(sp.strip()) for sp in subproceso_raw.split(',') if sp.strip()]

    if not matriz or not lista_subprocesos or not clase_doc or not archivos:
        flash("Todos los campos son obligatorios", "danger")
        return redirect(url_for('admin'))

    clase_doc_segura = secure_filename(clase_doc).replace("_", " ")
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Pre-leer archivos válidos en memoria (para poder guardarlos múltiples veces)
    archivos_validos = []
    invalidos = 0
    for archivo in archivos:
        if archivo.filename:
            filename = secure_filename(archivo.filename)
            if extension_permitida(filename):
                archivos_validos.append((filename, archivo.read()))
            else:
                invalidos += 1

    guardados = 0
    for sp in lista_subprocesos:
        ruta_destino = os.path.join(base_dir, '..', 'volumen_compartido', 'dataset_entrenamiento', matriz, sp, clase_doc_segura)
        os.makedirs(ruta_destino, exist_ok=True)
        
        for filename, datos in archivos_validos:
            ruta_archivo = os.path.join(ruta_destino, filename)
            with open(ruta_archivo, 'wb') as f:
                f.write(datos)
            guardados += 1

    subprocesos_str = ", ".join(lista_subprocesos)
    mensaje = f"✅ Éxito: Se guardaron {guardados} archivos en total para los subprocesos [{subprocesos_str}] en la categoría '{clase_doc_segura}'."
    if invalidos:
        mensaje += f" ({invalidos} archivo(s) inválido(s) omitido(s). Use TIF, PDF, JPG o PNG)."

    flash(mensaje, "success")
    return redirect(url_for('admin'))

@app.route('/admin/entrenar', methods=['POST'])
@login_requerido
def entrenar_modelos():
    if session['rol'] not in ('admin', 'superadmin'):
        return jsonify({"error": "No autorizado"}), 403
    
    set_estado('progreso_entrenamiento', '0')
    set_estado('entrenamiento', 'PROCESANDO')

    return jsonify({"mensaje": "Orden enviada correctamente"})

@app.route('/admin/estado_entrenamiento')
@limiter.exempt
@login_requerido
def estado_entrenamiento():
    estado = get_estado('entrenamiento', 'listo').lower()
    progreso = get_estado('progreso_entrenamiento', '0')
    try:
        progreso = int(progreso)
    except:
        progreso = 0
    return jsonify({"estado": estado, "progreso": progreso})

@app.route('/admin/refresh_inventario')
@login_requerido
def refresh_inventario():
    """Retorna el inventario y los modelos conocidos en JSON para actualización parcial del UI."""
    inventario = obtener_inventario_tipos_documentales()
    modelos = obtener_modelos_conocidos()
    return jsonify({
        'inventario': inventario,
        'modelos': modelos
    })

# ==========================================
# 6. RUTAS DE SUPERADMINISTRADOR
# ==========================================
@app.route('/superadmin')
@login_requerido
def superadmin():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado. Módulo exclusivo para Superadministradores.", "danger")
        return redirect(url_for('dashboard'))
    modelos_conocidos = obtener_modelos_conocidos()
    usuarios = listar_usuarios()
    return render_template(
        'superadmin.html',
        usuario=session['usuario'],
        rol=session['rol'],
        modelos_conocidos=modelos_conocidos,
        usuarios=usuarios
    )

@app.route('/superadmin/exportar_ia', methods=['GET'])
@login_requerido
def exportar_ia():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
    
    import shutil
    import time
    import tempfile
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    ruta_cerebros = os.path.join(base_dir, '..', 'volumen_compartido', 'cerebros_ia')
    ruta_processed = os.path.join(base_dir, '..', 'volumen_compartido', 'dataset_entrenamiento', 'processed')
    
    try:
        # Crear una carpeta temporal que contenga ambas fuentes de conocimiento
        tmp_export_dir = os.path.join(tempfile.gettempdir(), f'ia_export_{int(time.time())}')
        os.makedirs(tmp_export_dir, exist_ok=True)
        
        # 1. Copiar los cerebros compilados (modelo.pkl + vectorizador.pkl)
        if os.path.isdir(ruta_cerebros):
            shutil.copytree(ruta_cerebros, os.path.join(tmp_export_dir, 'cerebros_ia'))
        
        # 2. Copiar los datos históricos de entrenamiento (textos cacheados para reentrenar)
        if os.path.isdir(ruta_processed):
            shutil.copytree(ruta_processed, os.path.join(tmp_export_dir, 'processed'))
        
        ruta_zip_salida = os.path.join(tempfile.gettempdir(), f'cerebros_ia_export_{int(time.time())}')
        shutil.make_archive(ruta_zip_salida, 'zip', tmp_export_dir)
        archivo_final = ruta_zip_salida + '.zip'
        
        # Limpiar la carpeta temporal de montaje
        shutil.rmtree(tmp_export_dir, ignore_errors=True)
        
        logging.info(f"[SUPERADMIN] {session['usuario']} exportó los conocimientos de la IA (cerebros + datos históricos).")
        return send_file(archivo_final, as_attachment=True, download_name='Conocimiento_IA_Exportado.zip')
    except Exception as e:
        flash(f"Error al generar exportación: {e}", "danger")
        return redirect(url_for('superadmin'))

@app.route('/superadmin/importar_ia', methods=['POST'])
@login_requerido
def importar_ia():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
        
    archivo_zip = request.files.get('archivo_zip')
    if not archivo_zip or not archivo_zip.filename.endswith('.zip'):
        flash("Por favor sube un archivo .zip válido.", "danger")
        return redirect(url_for('superadmin'))
        
    import shutil
    import tempfile
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    ruta_cerebros = os.path.join(base_dir, '..', 'volumen_compartido', 'cerebros_ia')
    ruta_processed = os.path.join(base_dir, '..', 'volumen_compartido', 'dataset_entrenamiento', 'processed')
    os.makedirs(ruta_cerebros, exist_ok=True)
    os.makedirs(ruta_processed, exist_ok=True)
    
    try:
        import zipfile
        tmp_path = os.path.join(tempfile.gettempdir(), secure_filename(archivo_zip.filename))
        archivo_zip.save(tmp_path)
        
        # Validar el contenido del ZIP para prevenir Path Traversal / Zip Slip
        with zipfile.ZipFile(tmp_path, 'r') as zip_ref:
            for member in zip_ref.namelist():
                if member.startswith('/') or '..' in member:
                    raise ValueError(f"Archivo sospechoso en el ZIP (Posible Path Traversal): {member}")
            
            # Extraer a carpeta temporal para clasificar el contenido
            tmp_extract = os.path.join(tempfile.gettempdir(), f'ia_import_{os.getpid()}')
            zip_ref.extractall(tmp_extract)
        
        os.remove(tmp_path)
        
        # Detectar si es el formato nuevo (con subcarpetas cerebros_ia/ y processed/)
        ruta_cerebros_zip = os.path.join(tmp_extract, 'cerebros_ia')
        ruta_processed_zip = os.path.join(tmp_extract, 'processed')
        
        if os.path.isdir(ruta_cerebros_zip):
            # FORMATO NUEVO: El ZIP tiene carpetas separadas
            # 1. Copiar cerebros compilados
            for item in os.listdir(ruta_cerebros_zip):
                src = os.path.join(ruta_cerebros_zip, item)
                dst = os.path.join(ruta_cerebros, item)
                if os.path.isdir(src):
                    if os.path.exists(dst):
                        shutil.rmtree(dst)
                    shutil.copytree(src, dst)
                else:
                    shutil.copy2(src, dst)
            
            # 2. Copiar datos históricos de entrenamiento (para que el reentrenamiento los use)
            if os.path.isdir(ruta_processed_zip):
                for item in os.listdir(ruta_processed_zip):
                    src = os.path.join(ruta_processed_zip, item)
                    dst = os.path.join(ruta_processed, item)
                    if os.path.isdir(src):
                        # Fusionar: si la carpeta ya existe, copiar archivos nuevos sin borrar los existentes
                        if os.path.exists(dst):
                            for sub_item in os.listdir(src):
                                sub_src = os.path.join(src, sub_item)
                                sub_dst = os.path.join(dst, sub_item)
                                if os.path.isdir(sub_src):
                                    if not os.path.exists(sub_dst):
                                        shutil.copytree(sub_src, sub_dst)
                                    else:
                                        # Fusionar archivos individuales dentro de la clase
                                        for f in os.listdir(sub_src):
                                            shutil.copy2(os.path.join(sub_src, f), os.path.join(sub_dst, f))
                                else:
                                    shutil.copy2(sub_src, sub_dst)
                        else:
                            shutil.copytree(src, dst)
                    else:
                        shutil.copy2(src, dst)
                
                logging.info(f"[IMPORTAR] Datos históricos de entrenamiento restaurados en {ruta_processed}")
        else:
            # FORMATO ANTIGUO (retrocompatibilidad): El ZIP solo tenía los .pkl sueltos
            for item in os.listdir(tmp_extract):
                src = os.path.join(tmp_extract, item)
                dst = os.path.join(ruta_cerebros, item)
                if os.path.isdir(src):
                    if os.path.exists(dst):
                        shutil.rmtree(dst)
                    shutil.copytree(src, dst)
                else:
                    shutil.copy2(src, dst)
        
        # Limpiar carpeta temporal de extracción
        shutil.rmtree(tmp_extract, ignore_errors=True)
        
        flash("✅ Modelos de IA importados e instalados correctamente (cerebros + datos históricos).", "success")
        logging.info(f"[SUPERADMIN] {session['usuario']} importó un paquete de conocimiento IA (formato {'nuevo' if os.path.isdir(ruta_cerebros_zip) else 'legacy'}).")
    except Exception as e:
        flash(f"❌ Error al importar IA: {e}", "danger")
        logging.error(f"Error importando IA: {e}")
        
    return redirect(url_for('superadmin'))

@app.route('/superadmin/borrar_todo', methods=['POST'])
@login_requerido
def borrar_todo():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
    errores = borrar_todo_el_conocimiento()
    if errores:
        flash(f"⚠️ Borrado completado con advertencias: {'; '.join(errores)}", "warning")
    else:
        flash("✅ Todos los conocimientos de la IA han sido eliminados exitosamente.", "success")
    logging.info(f"[SUPERADMIN] {session['usuario']} ejecutó BORRADO TOTAL de conocimientos.")
    return redirect(url_for('superadmin'))

@app.route('/superadmin/borrar_proceso', methods=['POST'])
@login_requerido
def borrar_proceso():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
    matriz = secure_filename(request.form.get('matriz', '').strip())
    proceso = secure_filename(request.form.get('proceso', '').strip())
    if not matriz or not proceso:
        flash("Parámetros inválidos.", "danger")
        return redirect(url_for('superadmin'))
    errores = borrar_conocimiento_proceso(matriz, proceso)
    nombre_matriz = 'Natural' if matriz == 'BT' else ('Jurídico' if matriz == 'BR' else matriz)
    if errores:
        flash(f"⚠️ Proceso {proceso} ({nombre_matriz}) borrado con advertencias: {'; '.join(errores)}", "warning")
    else:
        flash(f"✅ Proceso {proceso} ({nombre_matriz}) eliminado exitosamente.", "success")
    logging.info(f"[SUPERADMIN] {session['usuario']} borró proceso {matriz}/{proceso}.")
    return redirect(url_for('superadmin'))

@app.route('/superadmin/borrar_clase', methods=['POST'])
@login_requerido
def borrar_clase():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
    matriz = request.form.get('matriz', '').strip().replace('/', '').replace('\\', '').replace('..', '')
    proceso = request.form.get('proceso', '').strip().replace('/', '').replace('\\', '').replace('..', '')
    clase = request.form.get('clase', '').strip().replace('/', '').replace('\\', '').replace('..', '')
    if not matriz or not proceso or not clase:
        flash("Parámetros inválidos.", "danger")
        return redirect(request.referrer or url_for('superadmin'))
    errores = borrar_conocimiento_clase(matriz, proceso, clase)
    nombre_matriz = 'Natural' if matriz == 'BT' else ('Jurídico' if matriz == 'BR' else matriz)
    
    # Disparar reentrenamiento automático para regenerar el cerebro sin esa clase
    try:
        from core.routing_ia import limpiar_cache
        limpiar_cache(matriz, proceso)  # Limpiar RAM para no usar cerebro viejo
        set_estado('progreso_entrenamiento', '0')
        set_estado('entrenamiento', 'PROCESANDO')
        logging.info(f"🔄 Reentrenamiento automático disparado tras borrar clase '{clase}' de {matriz}/{proceso}.")
    except Exception as e:
        logging.error(f"Error disparando reentrenamiento tras borrar clase: {e}")
    
    if errores:
        flash(f"⚠️ Clase '{clase}' del proceso {proceso} ({nombre_matriz}) borrada con advertencias: {'; '.join(errores)}", "warning")
    else:
        flash(f"✅ Clase '{clase}' del proceso {proceso} ({nombre_matriz}) eliminada. La IA se está reentrenando automáticamente.", "success")
    logging.info(f"[SUPERADMIN] {session['usuario']} borró clase {clase} de {matriz}/{proceso}.")
    return redirect(request.referrer or url_for('superadmin'))

@app.route('/admin/descartar_clase', methods=['POST'])
@login_requerido
def descartar_clase():
    if session['rol'] not in ['admin', 'superadmin']:
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
        
    matriz = request.form.get('matriz', '').strip().replace('/', '').replace('\\', '').replace('..', '')
    proceso = request.form.get('proceso', '').strip().replace('/', '').replace('\\', '').replace('..', '')
    clase = request.form.get('clase', '').strip().replace('/', '').replace('\\', '').replace('..', '')
    
    if not matriz or not proceso or not clase:
        flash("Parámetros inválidos.", "danger")
        return redirect(request.referrer or url_for('admin'))
        
    errores = descartar_subida_clase(matriz, proceso, clase)
    nombre_matriz = 'Natural' if matriz == 'BT' else ('Jurídico' if matriz == 'BR' else matriz)
    
    if errores:
        flash(f"⚠️ Hubo problemas al descartar la clase '{clase}': {'; '.join(errores)}", "warning")
    else:
        flash(f"✅ Subida de la clase '{clase}' cancelada y removida del inventario.", "success")
        
    logging.info(f"[{session['rol'].upper()}] {session['usuario']} descartó subida de la clase {clase} de {matriz}/{proceso}.")
    return redirect(request.referrer or url_for('admin'))

@app.route('/superadmin/crear_usuario', methods=['POST'])
@login_requerido
def crear_usuario_route():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
    username = request.form.get('nuevo_username', '').strip()
    password = request.form.get('nuevo_password', '').strip()
    rol = request.form.get('nuevo_rol', '').strip()
    exito, mensaje = crear_usuario(username, password, rol)
    if exito:
        flash(f"✅ {mensaje}", "success")
        logging.info(f"[SUPERADMIN] {session['usuario']} creó usuario '{username}' con rol '{rol}'.")
    else:
        flash(f"❌ {mensaje}", "danger")
    return redirect(url_for('superadmin'))

@app.route('/superadmin/eliminar_usuario', methods=['POST'])
@login_requerido
def eliminar_usuario_route():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
    user_id = request.form.get('user_id', '').strip()
    if not user_id:
        flash("ID de usuario inválido.", "danger")
        return redirect(url_for('superadmin'))
    exito, mensaje = eliminar_usuario(int(user_id), session['usuario'])
    if exito:
        flash(f"✅ {mensaje}", "success")
        logging.info(f"[SUPERADMIN] {session['usuario']} eliminó usuario ID {user_id}.")
    else:
        flash(f"❌ {mensaje}", "danger")
    return redirect(url_for('superadmin'))

@app.route('/superadmin/cambiar_clave', methods=['POST'])
@login_requerido
def cambiar_clave_route():
    if session['rol'] != 'superadmin':
        flash("Acceso denegado.", "danger")
        return redirect(url_for('dashboard'))
    user_id = request.form.get('user_id', '').strip()
    nueva_clave = request.form.get('nueva_clave', '').strip()
    if not user_id or not nueva_clave:
        flash("Datos incompletos.", "danger")
        return redirect(url_for('superadmin'))
    exito, mensaje = cambiar_clave_usuario(int(user_id), nueva_clave, session['usuario'], session['rol'])
    if exito:
        flash(f"✅ {mensaje}", "success")
        logging.info(f"[SUPERADMIN] {session['usuario']} cambió la clave del usuario ID {user_id}.")
    else:
        flash(f"❌ {mensaje}", "danger")
    return redirect(url_for('superadmin'))

@app.route('/api/reanudar_lote/<task_id>', methods=['POST'])
@login_requerido
def api_reanudar_lote(task_id):
    from core.auditor import procesar_lote_kofax_task
    # Volver a lanzar la misma tarea en Celery usando el ID existente
    procesar_lote_kofax_task.apply_async(args=[task_id], task_id=task_id)
    logging.info(f"Reanudando auditoría lote: {task_id}")
    return jsonify({"mensaje": "Auditoría reanudada en segundo plano", "task_id": task_id})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)